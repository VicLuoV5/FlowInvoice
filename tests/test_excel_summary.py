import importlib
import sys
import tempfile
import types
import unittest
from pathlib import Path


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


try:
    import pandas  # noqa: F401
    import openpyxl
except ImportError as exc:  # pragma: no cover - exercised only in minimal envs
    raise unittest.SkipTest(f"Excel dependencies unavailable: {exc}")


if "fitz" not in sys.modules:
    sys.modules["fitz"] = types.SimpleNamespace()
if "rapidocr_onnxruntime" not in sys.modules:
    sys.modules["rapidocr_onnxruntime"] = types.SimpleNamespace(RapidOCR=lambda: None)


processor = importlib.import_module("core.processor")


class ExcelSummaryTests(unittest.TestCase):
    def test_cover_sheet_summarizes_net_tax_and_total_by_existing_business_type(self):
        invoices = [
            _invoice("taxi-01.pdf", "打车/交通票", "100.00", "3.00", "103.00"),
            _invoice("taxi-02.pdf", "打车/交通票", "50.00", "1.50", "51.50"),
            _invoice("meal-01.pdf", "餐饮发票", "80.00", "4.80", "84.80"),
            _invoice("hotel-01.pdf", "住宿发票", "⚠️ 需手动核对", "⚠️ 需手动核对", "381.22"),
        ]

        with tempfile.TemporaryDirectory() as tmp:
            output = Path(tmp) / "summary.xlsx"

            success, message = processor.write_excel_from_data(invoices, str(output), "测试人")

            self.assertTrue(success, message)
            wb = openpyxl.load_workbook(output, data_only=False)
            ws = wb["封面汇总"]

            self.assertEqual(
                ["业务分类", "票据数量", "不含税金额合计", "税额合计", "报销金额合计（元）"],
                [ws.cell(row=8, column=col).value for col in range(1, 6)],
            )

            rows = {
                ws.cell(row=row, column=1).value: row
                for row in range(9, 13)
                if ws.cell(row=row, column=1).value
            }

            self.assertEqual(["打车/交通票", "餐饮发票", "住宿发票", "合  计"], list(rows))
            self.assertEqual(
                [
                    "=COUNTIF('报销明细'!$B:$B,A9)",
                    "=SUMIF('报销明细'!$B:$B,A9,'报销明细'!$E:$E)",
                    "=SUMIF('报销明细'!$B:$B,A9,'报销明细'!$F:$F)",
                    "=SUMIF('报销明细'!$B:$B,A9,'报销明细'!$G:$G)",
                ],
                [ws.cell(row=rows["打车/交通票"], column=col).value for col in range(2, 6)],
            )
            self.assertEqual(
                ["=SUM(B9:B11)", "=SUM(C9:C11)", "=SUM(D9:D11)", "=SUM(E9:E11)"],
                [ws.cell(row=rows["合  计"], column=col).value for col in range(2, 6)],
            )
            self.assertNotIn("飞机行程单", rows)
            self.assertEqual("auto", wb.calculation.calcMode)
            self.assertTrue(wb.calculation.fullCalcOnLoad)
            self.assertTrue(wb.calculation.forceFullCalc)


def _invoice(filename, inv_type, net, tax, total):
    return {
        "文件名": filename,
        "业务分类": inv_type,
        "日期": "2026年06月22日",
        "发票号码": filename.replace(".pdf", ""),
        "不含税金额": net,
        "税额": tax,
        "价税合计(报销额)": total,
        "置信度(%)": 95,
        "备注": "",
    }


if __name__ == "__main__":
    unittest.main()
