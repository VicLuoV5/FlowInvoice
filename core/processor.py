import os
import datetime
import fitz
import re
from collections import Counter
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidocr_onnxruntime import RapidOCR

import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# ================= 逻辑 A：智能合并排版 =================
def merge_pdfs_logic(input_folder, output_filename, layout_mode='横向', progress_callback=None):
    if os.path.exists(output_filename):
        try:
            os.remove(output_filename)
        except PermissionError:
            return False, f"❌ 文件正被占用！\n请先关闭正在打开的【{os.path.basename(output_filename)}】后再试。"

    merged_pdf = fitz.open()
    valid_exts = ('.pdf', '.jpg', '.jpeg', '.png')
    files = [os.path.basename(f) for f in os.listdir(input_folder) if f.lower().endswith(valid_exts)]
    files.sort()

    if not files:
        return False, "发票箱里空空如也，请先放入文件。"

    if layout_mode == '竖向':
        canvas_w, canvas_h = 595.0, 842.0
    else:
        canvas_w, canvas_h = 842.0, 595.0

    for idx, filename in enumerate(files, 1):
        if progress_callback:
            progress_callback(idx, len(files), filename)

        file_path = os.path.join(input_folder, filename)
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            img_doc = fitz.open(file_path)
            pdf_bytes = img_doc.convert_to_pdf()
            current_pdf = fitz.open("pdf", pdf_bytes)
        else:
            current_pdf = fitz.open(file_path)

        for page in current_pdf:
            new_page = merged_pdf.new_page(width=canvas_w, height=canvas_h)
            src_w, src_h = page.rect.width, page.rect.height

            rotation = 0
            effective_w, effective_h = src_w, src_h

            is_canvas_landscape = canvas_w > canvas_h
            is_source_landscape = src_w > src_h

            if is_canvas_landscape != is_source_landscape:
                rotation = 90
                effective_w, effective_h = src_h, src_w

            scale = min(canvas_w / effective_w, canvas_h / effective_h) * config.MERGE_SCALE_FACTOR

            new_w = effective_w * scale
            new_h = effective_h * scale
            x0 = (canvas_w - new_w) / 2
            y0 = (canvas_h - new_h) / 2
            target_rect = fitz.Rect(x0, y0, x0 + new_w, y0 + new_h)

            new_page.show_pdf_page(target_rect, current_pdf, page.number, rotate=rotation)

    try:
        merged_pdf.save(output_filename)
        merged_pdf.close()
        return True, f"成功合并 {len(files)} 份文件！"
    except Exception as e:
        return False, f"保存失败: {e}"


# ================= 逻辑 B：智能提取数据 =================

def _calc_confidence(date, num, total):
    """计算单张票据识别置信度 (0-100)。"""
    score = 0
    if date != "未抓取" and "模糊" not in date:
        score += 30
    elif date != "未抓取":
        score += 15
    if num != "未抓取":
        score += 20
    if total > 0:
        score += 50
    return score


def _classify_and_extract(filename, clean_text, nums):
    """对一段清洗后的文本进行分类 + 金额提取。失败返回 None。"""
    is_valid = False
    inv_type, total, net, tax = "未知票据", 0.0, 0.0, 0.0

    if "航空运输电子客票行程单" in clean_text or "机票" in clean_text or "航空客票" in clean_text:
        is_valid, inv_type = True, "机票行程单"

        m_t = re.search(r'(?:合计|金额|小写)[a-zA-Z¥￥]*(\d+\.\d{2})', clean_text)
        total = float(m_t.group(1)) if m_t else (nums[0] if nums else 0.0)

        m_tax = re.search(r'增值税税额[a-zA-Z¥￥]*(\d+\.\d{2})', clean_text)
        if m_tax:
            tax = float(m_tax.group(1))
            net = round(total - tax, 2)
        else:
            m_fund = re.search(r'民航发展基金[a-zA-Z¥￥]*(\d+\.\d{2})', clean_text)
            fund = float(m_fund.group(1)) if m_fund else 0.0
            taxable = total - fund
            if taxable > 0:
                net = round(taxable / config.FLIGHT_TAX_RATE, 2)
                tax = round(taxable - net, 2)
                net = round(total - tax, 2)
            else:
                net = round(total / config.FLIGHT_TAX_RATE, 2)
                tax = round(total - net, 2)

    elif "铁路电子客票" in clean_text or "高铁" in clean_text or "火车" in clean_text:
        is_valid, inv_type = True, "高铁/火车票"
        m_t = re.search(r'[¥￥]?(\d+\.\d{2})', clean_text)
        total = float(m_t.group(1)) if m_t else (nums[0] if nums else 0.0)
        net = round(total / config.HSR_TAX_RATE, 2)
        tax = round(total - net, 2)

    elif any(k in clean_text for k in ["出租", "打车", "运输服务", "上车", "下车", "里程", "等候", "机打发票"]):
        is_valid, inv_type = True, "打车/交通票"

        m_yuan = re.search(r'(\d{1,4}(?:[\.,]\d{1,2})?)[元]', clean_text)
        m_taxi = re.search(r'(?:总额|实收|附加|金额|应收)[^\d]{0,6}(\d{1,4}(?:[\.,]\d{1,2})?)', clean_text)

        if m_yuan:
            total = float(m_yuan.group(1).replace(',', '.'))
        elif m_taxi:
            total = float(m_taxi.group(1).replace(',', '.'))
        else:
            valid_taxi_nums = [n for n in nums if n < config.MAX_TAXI_AMOUNT]
            if valid_taxi_nums:
                total = valid_taxi_nums[0]

        if total >= config.MAX_TAXI_AMOUNT:
            total = 0.0

        if total > 0:
            net = round(total / config.TAXI_TAX_RATE, 2)
            tax = round(total - net, 2)

    elif any(k in clean_text for k in ["中国石油", "中国石化", "加油站", "成品油", "汽油", "柴油"]):
        is_valid, inv_type = True, "加油费"

    elif any(k in clean_text for k in ["中国移动", "中国联通", "中国电信", "话费", "通讯费发票"]):
        is_valid, inv_type = True, "通讯费"

    elif "餐饮" in clean_text:
        is_valid, inv_type = True, "餐饮发票"

    elif "住宿" in clean_text or "酒店" in clean_text:
        is_valid, inv_type = True, "住宿发票"

    elif "发票" in clean_text:
        is_valid, inv_type = True, "增值税发票"

    # 增值税发票类通用算法（金额配对）
    VAT_TYPES = {"餐饮发票", "住宿发票", "增值税发票", "加油费", "通讯费"}
    if is_valid and inv_type in VAT_TYPES:
        m_t = re.search(r'(?:价税合计|小写|总计)[a-zA-Z¥￥]*(\d+\.\d{2})', clean_text)
        if m_t:
            total = float(m_t.group(1))
        elif nums:
            total = nums[0]

        if total > 0:
            pool = nums + [0.00] if 0.00 not in nums else nums
            for i in range(len(pool)):
                for j in range(i, len(pool)):
                    if abs((pool[i] + pool[j]) - total) <= 0.02:
                        net, tax = max(pool[i], pool[j]), min(pool[i], pool[j])
                        break

    if not is_valid:
        return None
    return inv_type, total, net, tax


def extract_invoices_data(input_folder, progress_callback=None):
    """提取 input_folder 下所有发票的结构化数据，不写 Excel。

    Returns:
        (invoices, failures)
        invoices: List[dict] — 每张票的结构化字段
        failures: List[dict] — 形如 {"file": 文件名, "reason": 原因}
    """
    valid_exts = ('.pdf', '.jpg', '.jpeg', '.png')
    files = [os.path.basename(f) for f in os.listdir(input_folder) if f.lower().endswith(valid_exts)]

    if not files:
        return [], []

    try:
        ocr_engine = RapidOCR()
    except Exception as e:
        return [], [{"file": "(OCR 初始化)", "reason": str(e)}]

    all_invoices = []
    failures = []

    for idx, filename in enumerate(files, 1):
        if progress_callback:
            progress_callback(idx, len(files), filename)

        file_path = os.path.join(input_folder, filename)
        raw_text = ""
        try:
            if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                ocr_result, _ = ocr_engine(file_path)
                if ocr_result:
                    raw_text = " ".join([item[1] for item in ocr_result])
            else:
                doc = fitz.open(file_path)
                for page in doc:
                    raw_text += page.get_text("text").replace('\n', ' ') + " "
                doc.close()
        except Exception as e:
            failures.append({"file": filename, "reason": f"读取失败：{e}"})
            continue

        if not raw_text.strip():
            failures.append({"file": filename, "reason": "OCR 未识别到文本（图片可能过于模糊）"})
            continue

        clean_text = raw_text.replace(" ", "")

        # 日期
        date = "未抓取"
        m_date = re.search(r'(20\d{2})[年\-./](\d{1,2})[月\-./](\d{1,2})[日]?', clean_text)
        if m_date:
            y, m, d = m_date.groups()
            date = f"{y}年{m.zfill(2)}月{d.zfill(2)}日"
        else:
            m_date2 = re.search(r'(20\d{2})[年\-./](\d{1,2})[月]?', clean_text)
            if m_date2:
                y, m = m_date2.groups()
                date = f"{y}年{m.zfill(2)}月(日模糊)"

        # 发票号码
        num = "未抓取"
        m_num = re.search(r'(?:发票号码|号码|No)[：:]?(\d{8,24})', clean_text, re.IGNORECASE)
        if m_num:
            num = m_num.group(1)
        else:
            m_num_fallback = re.search(r'(?<!\d)(\d{8})(?!\d)', clean_text)
            if m_num_fallback:
                num = m_num_fallback.group(1)

        # 数字池
        nums = [float(x) for x in re.findall(r'\d+\.\d{2}', clean_text)]
        nums = sorted(list(set([a for a in nums if a < 1000000])), reverse=True)

        # 分类 + 金额提取
        result = _classify_and_extract(filename, clean_text, nums)
        if result is None:
            failures.append({"file": filename, "reason": "无法识别票据类型（无匹配关键词）"})
            continue

        inv_type, total, net, tax = result

        str_total = f"{total:.2f}" if total > 0 else "⚠️ 需手动核对"
        str_net   = f"{net:.2f}"   if total > 0 else "⚠️ 需手动核对"
        str_tax   = f"{tax:.2f}"   if total > 0 else "⚠️ 需手动核对"

        remark = ""
        if inv_type == "餐饮发票":
            remark = "餐饮税额不可抵扣"
        elif total == 0:
            remark = "字迹太模糊，请人工核票"

        all_invoices.append({
            "文件名": filename,
            "业务分类": inv_type,
            "日期": date,
            "发票号码": num,
            "不含税金额": str_net,
            "税额": str_tax,
            "价税合计(报销额)": str_total,
            "置信度(%)": _calc_confidence(date, num, total),
            "备注": remark,
        })

    # 重复发票号检测
    num_counts = Counter(inv["发票号码"] for inv in all_invoices if inv["发票号码"] != "未抓取")
    for inv in all_invoices:
        if num_counts.get(inv["发票号码"], 0) > 1:
            dup_note = "⚠️ 疑似重复"
            inv["备注"] = f"{inv['备注']} · {dup_note}" if inv["备注"] else dup_note

    return all_invoices, failures


def write_excel_from_data(invoices, output_excel, submitter_name=""):
    """从已提取的发票数据写入 Excel 文件。"""
    if os.path.exists(output_excel):
        try:
            os.remove(output_excel)
        except PermissionError:
            return False, f"❌ 文件正被占用！\n请先关闭正在打开的【{os.path.basename(output_excel)}】后再试。"

    if not invoices:
        return False, "无数据可写入。"

    _write_excel(invoices, output_excel, submitter_name)
    return True, f"Excel 已生成，共 {len(invoices)} 张票据"


def extract_data_logic(input_folder, output_excel, submitter_name="", progress_callback=None):
    """便捷函数：一步完成 提取 + 写 Excel（供桌面端使用，保持向后兼容）。"""
    if os.path.exists(output_excel):
        try:
            os.remove(output_excel)
        except PermissionError:
            return False, f"❌ 文件正被占用！\n请先关闭正在打开的【{os.path.basename(output_excel)}】后再试。"

    invoices, failures = extract_invoices_data(input_folder, progress_callback)

    if not invoices:
        if failures:
            reasons = "\n".join(f"• {f['file']}：{f['reason']}" for f in failures[:5])
            tail = f"\n（还有 {len(failures) - 5} 个...）" if len(failures) > 5 else ""
            return False, f"未提取到任何有效数据。\n\n失败详情：\n{reasons}{tail}"
        return False, "未找到有效文件。"

    _write_excel(invoices, output_excel, submitter_name)

    msg = f"提取成功，共 {len(invoices)} 张票据"
    if failures:
        msg += f"\n\n⚠️ {len(failures)} 个文件未能识别："
        for f in failures[:5]:
            msg += f"\n• {f['file']}：{f['reason']}"
        if len(failures) > 5:
            msg += f"\n（还有 {len(failures) - 5} 个...）"
    return True, msg


def _write_excel(all_invoices, output_excel, submitter_name):
    """生成带封面汇总页和置信度条件格式的 Excel 报告。"""
    BLUE        = "4285F4"
    LIGHT_BLUE  = "E8F0FE"
    LIGHT_GRAY  = "F8F9FA"
    RED_FILL    = "FFDDD2"
    YELLOW_FILL = "FFF8E1"
    thin        = Side(style='thin', color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    df = pd.DataFrame(all_invoices)
    writer = pd.ExcelWriter(output_excel, engine='openpyxl')
    wb = writer.book

    # ── Sheet 1: 封面汇总 ─────────────────────────────────────
    ws = wb.create_sheet("封面汇总", 0)

    def sc(row, col, value, bold=False, size=11, fg=None, bg=None, align='left', italic=False):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name='Microsoft YaHei', size=size, bold=bold,
                      color=fg or "000000", italic=italic)
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center')
        return c

    ws.merge_cells('A1:G1')
    sc(1, 1, config.APP_NAME, bold=True, size=18, fg="FFFFFF", bg=BLUE, align='center')
    ws.row_dimensions[1].height = 38

    ws.merge_cells('A2:G2')
    sc(2, 1, config.APP_SUBTITLE, size=11, fg="555555", bg="EAF0FF", align='center')
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 10

    sc(4, 1, "填报人", bold=True)
    sc(4, 2, submitter_name if submitter_name else "（请填写）")
    sc(4, 4, "生成日期", bold=True)
    sc(4, 5, datetime.date.today().strftime("%Y年%m月%d日"))
    ws.row_dimensions[4].height = 22

    ws.row_dimensions[5].height = 10

    ws.merge_cells('A6:G6')
    sc(6, 1, "—  报销分类汇总  —", bold=True, size=12, fg=BLUE, align='center')
    ws.row_dimensions[6].height = 26

    for col, hdr in enumerate(["业务分类", "票据数量", "报销金额合计（元）"], 1):
        c = sc(7, col, hdr, bold=True, fg="FFFFFF", bg=BLUE, align='center')
        c.border = cell_border
    ws.row_dimensions[7].height = 22

    type_summary = {}
    for inv in all_invoices:
        t = inv["业务分类"]
        try:
            amt = float(inv["价税合计(报销额)"])
        except (ValueError, TypeError):
            amt = 0.0
        entry = type_summary.setdefault(t, {"count": 0, "total": 0.0})
        entry["count"] += 1
        entry["total"] += amt

    data_row = 8
    grand_total, grand_count = 0.0, 0
    for inv_type, data in sorted(type_summary.items()):
        sc(data_row, 1, inv_type).border = cell_border
        c2 = sc(data_row, 2, data["count"], align='center')
        c2.border = cell_border
        c3 = ws.cell(row=data_row, column=3, value=round(data["total"], 2))
        c3.number_format = '#,##0.00'
        c3.border = cell_border
        c3.alignment = Alignment(horizontal='right', vertical='center')
        grand_total += data["total"]
        grand_count += data["count"]
        data_row += 1

    for col in range(1, 4):
        c = ws.cell(row=data_row, column=col)
        c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        c.border = cell_border
        c.font = Font(name='Microsoft YaHei', bold=True)
        c.alignment = Alignment(vertical='center',
                                horizontal='center' if col == 2 else ('right' if col == 3 else 'left'))
    ws.cell(row=data_row, column=1).value = "合  计"
    ws.cell(row=data_row, column=2).value = grand_count
    c_tot = ws.cell(row=data_row, column=3)
    c_tot.value = round(grand_total, 2)
    c_tot.number_format = '#,##0.00'
    ws.row_dimensions[data_row].height = 24

    if any(inv["业务分类"] == "餐饮发票" for inv in all_invoices):
        note_row = data_row + 2
        ws.merge_cells(f'A{note_row}:G{note_row}')
        c = ws.cell(row=note_row, column=1,
                    value="※ 餐饮票税额不可抵扣，报销额为价税合计，详见明细页备注。")
        c.font = Font(name='Microsoft YaHei', color="EA4335", italic=True, size=10)

    for col, w in zip('ABCDEFG', [20, 16, 22, 12, 18, 12, 12]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: 报销明细 ─────────────────────────────────────
    DISPLAY_COLS = ["文件名", "业务分类", "日期", "发票号码",
                    "不含税金额", "税额", "价税合计(报销额)", "置信度(%)", "备注"]
    df.to_excel(writer, index=False, sheet_name='报销明细', columns=DISPLAY_COLS)
    ws_data = writer.sheets['报销明细']

    hdr_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    for col_idx in range(1, len(DISPLAY_COLS) + 1):
        c = ws_data.cell(row=1, column=col_idx)
        c.fill = hdr_fill
        c.font = Font(name='Microsoft YaHei', bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws_data.row_dimensions[1].height = 22

    for row_idx, inv in enumerate(all_invoices, start=2):
        conf = inv.get("置信度(%)")
        if conf is None:
            conf = 100
        if conf < config.CONFIDENCE_LOW_THRESHOLD:
            row_fill = PatternFill("solid", fgColor=RED_FILL)
        elif conf < config.CONFIDENCE_HIGH_THRESHOLD:
            row_fill = PatternFill("solid", fgColor=YELLOW_FILL)
        else:
            row_fill = None
        if row_fill:
            for col_idx in range(1, len(DISPLAY_COLS) + 1):
                ws_data.cell(row=row_idx, column=col_idx).fill = row_fill

    for col_idx, w in enumerate([28, 14, 18, 22, 14, 12, 18, 10, 24], 1):
        ws_data.column_dimensions[get_column_letter(col_idx)].width = w

    writer.close()
